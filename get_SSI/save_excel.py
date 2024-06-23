import pandas as pd
import requests
from pandas import json_normalize
import numpy as np
from io import BytesIO
import time
from datetime import datetime, timedelta
import json
import polars as pl


idx = pd.IndexSlice
pd.options.display.float_format = '{:,.3f}'.format


# API request config for SSI API endpoints
ssi_headers = {
        'Connection': 'keep-alive',
        'sec-ch-ua': '"Not A;Brand";v="99", "Chromium";v="98", "Google Chrome";v="98"',
        'DNT': '1',
        'sec-ch-ua-mobile': '?0',
        'X-Fiin-Key': 'KEY',
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'X-Fiin-User-ID': 'ID',
        'X-Fiin-User-Token':'79,14,109,2,77,63,182,226,168,166,146,27,233,245,5,195,0,200,169,49,153,168,150,254,116,248,73,45,23,61,239,77,183,194,66,44,130,28,239,66,27,56,75,64,56,106,32,203,36,15,41,203,250,254,180,226,198,77,152,213,125,234,19,189,68,11,105,241,12,164,235,22,207,252,226,142,142,239,234,206,6,141,63,201,33,169,127,101,9,25,210,180,120,123,145,86,239,76,144,170,24,236,42,224,97,18,41,250,154,193,169,100,133,44,162,88,201,178,212,112,53,114,12,95,81,58,144,233,119,10,24,138,241,254,39,124,10,203,168,121,230,82,103,254,250,60,20,249,237,230,12,76,78,65,234,48,255,62,102,6,201,137,37,132,182,151,246,26,37,81,123,183,82,209,104,37,195,103,138,251,132,30,158,210,151,131,145,67,235,171,34,119,235,73,193,80,232,193,102,189,47,70,108,12,43,253,60,16,114,169,91,125,8,129,40,14,219,191,133,180,200,249,65,81,108,254,163,47,151,3,1,35,132,124,160,200,104,243,64,185,149,215,77,222,174,177,231,227,14,240,127,167,68,133,199,216',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36',
        'X-Fiin-Seed': 'SEED',
        'sec-ch-ua-platform': 'Windows',
        'Origin': 'https://iboard.ssi.com.vn',
        'Sec-Fetch-Site': 'same-site',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
        'Referer': 'https://iboard.ssi.com.vn/',
        'Accept-Language': 'en-US,en;q=0.9,vi-VN;q=0.8,vi;q=0.7'
        }


def organ_listing (lang='vi', headers=ssi_headers):
    """
    Return a DataFrame of all available stock symbols. Live data is retrieved from the SSI API.
    Parameters:
        lang (str): language of the data. Default is 'vi', other options are 'en'
        headers (dict): headers of the request
    """
    url = f"https://fiin-core.ssi.com.vn/Master/GetListOrganization?language={lang}"
    response = requests.request("GET", url, headers=headers)
    status = response.status_code
    if status == 200:
        data = response.json()
        # print('Total number of companies: ', data['totalCount'])
        df = pd.DataFrame(data['items'])
        return df
    else:
        print('Error in API response', response.text)
def financial_report (symbol='SSI', report_type='BalanceSheet', frequency='Quarterly', periods=200, latest_year=None, headers=ssi_headers): # Quarterly, Yearly
    """
    Return financial reports of a stock symbol by type and period.
    Args:
        symbol (:obj:`str`, required): 3 digits name of the desired stock.
        report_type (:obj:`str`, required): BalanceSheet, IncomeStatement, CashFlow
        report_range (:obj:`str`, required): Yearly or Quarterly.
    """
    symbol = symbol.upper()
    organ_code = organ_listing().query(f'ticker == @symbol')['organCode'].values[0]
    this_year = str(datetime.now().year)
    if latest_year == None:
      latest_year = this_year
    else:
      if isinstance(latest_year, int) != True:
        print('Please input latest_year as int number')
      else:
        pass
    url = f'https://fiin-fundamental.ssi.com.vn/FinancialStatement/Download{report_type}?language=vi&OrganCode={organ_code}&Skip=0&Frequency={frequency}&numberOfPeriod={periods}&latestYear={latest_year}'
    response = requests.get(url, headers=headers)
    # print(response.text)
    status = response.status_code
    if status == 200:
        df = pd.read_excel(BytesIO(response.content), skiprows=7)
        # .dropna()
        return df
    else:
        print(f'Error {status} when getting data from SSI. Details:\n {response.text}')
        return None

def mc(symbol='SSI',frequency='Quarterly'):
        headers = {
                'Connection': 'keep-alive',
                'sec-ch-ua': '"Not A;Brand";v="99", "Chromium";v="98", "Google Chrome";v="98"',
                'DNT': '1',
                'sec-ch-ua-mobile': '?0',
                'X-Fiin-Key': 'KEY',
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'X-Fiin-User-ID': 'ID',
                'X-Fiin-Key':'KEY',
                'X-Fiin-Seed':'SEED',
                'X-Fiin-User-Token':'93,203,163,40,224,188,115,115,138,126,18,199,199,124,39,108,231,125,80,15,79,226,178,184,60,101,162,174,35,156,160,54,113,153,99,49,167,98,81,217,225,67,146,16,255,228,25,242,213,192,129,186,139,181,191,112,119,41,36,49,45,37,208,216,184,215,157,52,95,29,185,63,186,228,97,27,86,163,49,131,67,17,92,172,156,132,217,88,15,231,7,175,164,138,29,180,116,130,76,38,107,88,132,186,75,8,124,209,185,88,180,7,211,235,229,42,232,206,219,25,84,76,226,0,197,66,181,79,230,74,208,200,86,229,25,9,26,44,219,167,162,161,178,144,90,239,165,36,41,99,186,205,217,181,7,162,101,238,186,34,56,31,153,19,176,193,110,47,18,237,192,133,113,67,194,227,13,202,239,126,23,189,121,36,77,74,211,188,203,144,113,209,48,248,84,22,237,223,154,232,228,74,124,239,104,86,146,26,237,250,25,53,58,197,214,59,195,119,6,146,131,42,111,254,70,220,1,233,163,205,29,132,94,248,229,177,175,42,32,130,189,20,83,218,153,0,143,234,246,130,103,169,144,137,128,169',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36',
                'X-Fiin-Seed': 'SEED',
                'sec-ch-ua-platform': 'Windows',
                'Origin': 'https://iboard.ssi.com.vn',
                'Sec-Fetch-Site': 'same-site',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Dest': 'empty',
                'Referer': 'https://iboard.ssi.com.vn/',
                'Accept-Language': 'en-US,en;q=0.9,vi-VN;q=0.8,vi;q=0.7'
                }
        symbol = symbol.upper()
        organ_code = organ_listing().query(f'ticker == @symbol')['organCode'].values[0]
        url = f'https://fiin-fundamental.ssi.com.vn/FinancialAnalysis/DownloadFinancialRatio2?language=vi&OrganCode={organ_code}&CompareToIndustry=false&Frequency={frequency}&Ratios=ryd11&TimeLineFrom=2001&TimeLineTo=2025'
        response = requests.get(url, headers=headers)
        status = response.status_code
        if status == 200:
            df = pd.read_excel(BytesIO(response.content), skiprows=7)
            df = df.rename(columns={'Chỉ số': 'LengthReport'})
            return df.iloc[[1],:]
        else:
            print(f'Error {status} when getting data from SSI. Details:\n {response.text}')
            return None



def get_mc_Y(ticker):    
    x = mc(ticker,'Yearly')
    x = x.T
    x.columns = x.iloc[0]
    x = x.iloc[1:,:]
    x = x.loc[~x.index.str.contains(r'\.')]
    x['dates'] = x.index.astype(int)
    #replace 0 with NaN value
    x = x.replace(0, np.nan)
    #rename column 2 to 'mc'
    x = x.rename(columns={x.columns[0]: 'mc'})
    return x

def get_mc_Q(ticker):
    x = mc(ticker,'Quarterly')
    x = x.T
    x.columns = x.iloc[0]
    x = x.iloc[1:,:]
    x['year'] = x.index.str[-4:].astype(int)
    x['quarter'] = x.index.str[1].astype(float)
    x['dates'] = pd.PeriodIndex(year=x["year"], quarter=x["quarter"])
    x['dates'] = x['dates'].dt.to_timestamp(freq='Q')
    x = x.replace(0, np.nan)
    x = x.rename(columns={x.columns[0]: 'mc'})
    x = x.sort_values(by='dates')
    return x


list_chitieu = ['dates','mc','TỔNG TÀI SẢN','TÀI SẢN NGẮN HẠN','Tiền và tương đương tiền','Giá trị thuần đầu tư ngắn hạn','Các khoản phải thu','Hàng tồn kho, ròng','TÀI SẢN DÀI HẠN','Phải thu dài hạn','Tài sản cố định','GTCL TSCĐ hữu hình','Nguyên giá TSCĐ hữu hình','Khấu hao lũy kế TSCĐ hữu hình','GTCL Tài sản thuê tài chính','Nguyên giá tài sản thuê tài chính','Khấu hao lũy kế tài sản thuê tài chính','GTCL tài sản cố định vô hình','Nguyên giá TSCĐ vô hình','Khấu hao lũy kế TSCĐ vô hình','Bất động sản đầu tư','Nguyên giá tài sản đầu tư','Khấu hao lũy kế tài sản đầu tư','Tài sản dở dang dài hạn','Đầu tư dài hạn',
                'NỢ PHẢI TRẢ','Nợ ngắn hạn','Phải trả người bán','Người mua trả tiền trước','Doanh thu chưa thực hiện ngắn hạn','Vay ngắn hạn','Nợ dài hạn','Người mua trả tiền trước dài hạn','Doanh thu chưa thực hiên','Vay dài hạn','Trái phiếu chuyển đổi','VỐN CHỦ SỞ HỮU','Vốn góp','Thặng dư vốn cổ phần','Cổ phiếu Quỹ','Lãi chưa phân phối','Lợi ích cổ đông không kiểm soát','Doanh số thuần','Lãi gộp','Thu nhập tài chính','Chi phí tài chính','Trong đó: Chi phí lãi vay','Lãi/(lỗ) từ công ty liên doanh','Chi phí bán hàng','Chi phí quản lý doanh  nghiệp','Thu nhập khác, ròng','Lãi/(lỗ) ròng trước thuế','Lãi/(lỗ) thuần sau thuế','Lợi nhuận của Cổ đông của Công ty mẹ',
                'Lưu chuyển tiền thuần từ các hoạt động sản xuất kinh doanh','Khấu hao TSCĐ',
                'Chi phí dự phòng','Chi phí lãi vay','Chi phí lãi vay đã trả','Thuế thu nhập doanh nghiệp đã trả',
                'Lưu chuyển tiền tệ ròng từ hoạt động đầu tư','Tiền mua tài sản cố định và các tài sản dài hạn khác',
                'Tiền thu được từ thanh lý tài sản cố định','Cổ tức và tiền lãi nhận được',
                'Lưu chuyển tiền tệ từ hoạt động tài chính','Tiền thu từ phát hành cổ phiếu và vốn góp',
                'Chi trả cho việc mua lại, trả lại cổ phiếu','Tiền thu được các khoản đi vay',
                'Tiển trả các khoản đi vay','Tiền thanh toán vốn gốc đi thuê tài chính','Cổ tức đã trả'
]



def add_ratios(x):
    x = x.select(list_chitieu)
    x = x.with_columns([
    (pl.col('Tiền và tương đương tiền') + pl.col('Giá trị thuần đầu tư ngắn hạn')).alias('bs_cash'),
    (pl.col('Các khoản phải thu') + pl.col('Phải thu dài hạn')).alias('bs_ar'),
    (pl.col('GTCL TSCĐ hữu hình') + pl.col('GTCL Tài sản thuê tài chính') + pl.col('GTCL tài sản cố định vô hình') + pl.col('Bất động sản đầu tư')).alias('bs_fa'),
    (pl.col('Nguyên giá TSCĐ hữu hình') + pl.col('Nguyên giá tài sản thuê tài chính') + pl.col('Nguyên giá TSCĐ vô hình') + pl.col('Nguyên giá tài sản đầu tư')).alias('bs_gross_fa'),
    (pl.col('TỔNG TÀI SẢN') - pl.col('Tiền và tương đương tiền') - pl.col('Giá trị thuần đầu tư ngắn hạn') - pl.col('Các khoản phải thu') - pl.col('Phải thu dài hạn') - pl.col('Hàng tồn kho, ròng') - (pl.col('GTCL TSCĐ hữu hình') + pl.col('GTCL Tài sản thuê tài chính') + pl.col('GTCL tài sản cố định vô hình') + pl.col('Bất động sản đầu tư')) - pl.col('Tài sản dở dang dài hạn') - pl.col('Đầu tư dài hạn')).alias('other_asset'),
    (pl.col('Doanh thu chưa thực hiện ngắn hạn') + pl.col('Doanh thu chưa thực hiên') + pl.col('Người mua trả tiền trước') + pl.col('Người mua trả tiền trước dài hạn')).alias('bs_cust_pre'),
    (pl.col('Vay ngắn hạn') + pl.col('Vay dài hạn') + pl.col('Trái phiếu chuyển đổi')).alias('debt'),
    (pl.col('NỢ PHẢI TRẢ') - pl.col('Vay ngắn hạn') - pl.col('Vay dài hạn') - pl.col('Trái phiếu chuyển đổi') - pl.col('Phải trả người bán') - (pl.col('Doanh thu chưa thực hiện ngắn hạn') + pl.col('Doanh thu chưa thực hiên') + pl.col('Người mua trả tiền trước') + pl.col('Người mua trả tiền trước dài hạn'))).alias('other_lia'),
    (pl.col('VỐN CHỦ SỞ HỮU') - pl.col('Vốn góp') - pl.col('Lãi chưa phân phối') - pl.col('Cổ phiếu Quỹ')).alias('other_equity'),
    ((pl.col('Vay ngắn hạn') + pl.col('Vay dài hạn') + pl.col('Trái phiếu chuyển đổi')) - (pl.col('Tiền và tương đương tiền') + pl.col('Giá trị thuần đầu tư ngắn hạn'))).alias('netdebt'),
    (pl.col('TÀI SẢN NGẮN HẠN') / (1 + pl.col('TỔNG TÀI SẢN'))).alias('ca/ta'),
    ((pl.col('Vay ngắn hạn') + pl.col('Vay dài hạn') + pl.col('Trái phiếu chuyển đổi')) / (1 + pl.col('VỐN CHỦ SỞ HỮU'))).alias('de'),
    (1 - pl.col('Lãi/(lỗ) thuần sau thuế') / (1 + pl.col('Lãi/(lỗ) ròng trước thuế'))).alias('tax_rate'),
    
    (pl.col('Lãi gộp') + pl.col('Chi phí bán hàng') + pl.col('Chi phí quản lý doanh  nghiệp')).alias('op'),
    pl.col('Khấu hao TSCĐ').alias('cf_dep'),
    (pl.col('Cổ tức đã trả') + pl.col('Chi trả cho việc mua lại, trả lại cổ phiếu')).alias('cf_div'),
    (pl.col('Tiền thu được các khoản đi vay') + pl.col('Tiển trả các khoản đi vay') + pl.col('Tiền thanh toán vốn gốc đi thuê tài chính')).alias('cf_delta_debt')
    ])
    x = x.with_columns([
    (pl.col('op') * (1 - pl.col('tax_rate'))).alias('core_e'),
    (pl.col('Thu nhập tài chính') + (pl.col('Chi phí tài chính') - pl.col('Trong đó: Chi phí lãi vay'))).alias('fin_income'),
    (pl.col('op') + pl.col('Trong đó: Chi phí lãi vay')).alias('EBT'),    
    ((pl.col('Lưu chuyển tiền thuần từ các hoạt động sản xuất kinh doanh') + pl.col('Lưu chuyển tiền tệ ròng từ hoạt động đầu tư') + pl.col('Lưu chuyển tiền tệ từ hoạt động tài chính')) - 
     (pl.col('Lãi/(lỗ) thuần sau thuế') + pl.col('Khấu hao TSCĐ') + pl.col('Tiền mua tài sản cố định và các tài sản dài hạn khác') + pl.col('Tiền thu từ phát hành cổ phiếu và vốn góp') + pl.col('cf_delta_debt') + pl.col('cf_div'))).alias('cf_khac'),
    (pl.col('op') + pl.col('cf_dep')).alias('operating_EBITDA'),
    ((pl.col('Lãi/(lỗ) ròng trước thuế') - pl.col('Thu nhập khác, ròng') + pl.col('cf_dep') - pl.col('Trong đó: Chi phí lãi vay')).alias('EBITDA'))
])

    return x

col4 = ['Lãi gộp', 'op', 'EBT', 'Lãi/(lỗ) ròng trước thuế', 'Lãi/(lỗ) thuần sau thuế', 'Lợi nhuận của Cổ đông của Công ty mẹ', 'core_e', 'EBITDA']
col5 = ['Doanh số thuần', 'Lãi gộp', 'op', 'EBT', 'Lãi/(lỗ) ròng trước thuế', 'Lãi/(lỗ) thuần sau thuế', 'Lợi nhuận của Cổ đông của Công ty mẹ', 'core_e', 'EBITDA']

def margin_func(x):
    for i in col4:
        margin_col = i + "_m"
        x = x.with_columns([(pl.col(i)/(1+pl.col('Doanh số thuần'))).alias(margin_col)])
    return x


def g_func(x):
    for i in col5:
        g_col = "g_" + i 
        x = x.with_columns((pl.col(i).pct_change(n=1)).alias(g_col))
    
    
    x = x.with_columns([
        (pl.col('Lãi/(lỗ) thuần sau thuế') / pl.col('VỐN CHỦ SỞ HỮU').rolling_mean(2, min_periods=2)).alias('roe'),
        (pl.col('core_e') / pl.col('VỐN CHỦ SỞ HỮU').rolling_mean(2, min_periods=2)).alias('roe_core'),
        (pl.col('Lãi/(lỗ) thuần sau thuế') / pl.col('TỔNG TÀI SẢN').rolling_mean(2, min_periods=2)).alias('roa')
    ])
    
    return x




def get_fs_Y(ticker):
    bs = financial_report(ticker,'BalanceSheet','Yearly')
    # bs = bs.loc[:, (bs==0).mean() < .6]
    pl = financial_report(ticker,'IncomeStatement','Yearly')
    # pl = pl.loc[:, (pl==0).mean() < .6]
    cf = financial_report(ticker,'CashFlow','Yearly')
    cf = cf.rename(columns={'Unnamed: 0': 'CHỈ TIÊU'})
    cf.set_index('CHỈ TIÊU', inplace=True)
    cf2 = pd.DataFrame(index = {'Khấu hao TSCĐ':'CHỈ TIÊU'}, columns = cf.columns).fillna(0)
    if cf2.index[0] in cf.index:
        pass
    else:
        cf = pd.concat([cf,cf2],axis=0)
    cf = cf.reset_index().rename(columns={'index': 'CHỈ TIÊU'})

    fs = pd.concat([bs,pl,cf])
    fs = fs.T
    fs.columns = fs.iloc[0]
    fs = fs.iloc[1:,:]
    #delete all the row with NaN value > 40
    fs = fs.dropna(axis=0,thresh=40)
    fs['dates'] = fs.index.astype(int)
    fs = fs.loc[:,~fs.columns.duplicated(keep='first')]
    try:
        mc = get_mc_Y(ticker)
    except:
        mc = pd.DataFrame(columns=['mc','dates'])
    fs = fs.merge(mc[['mc','dates']], on='dates', how='left')

    
    return fs
def get_data_Y(ticker):
    fs = get_fs_Y(ticker)
    fs = pl.from_pandas(fs)
    fs = add_ratios(fs)
    fs = margin_func(fs)
    fs = g_func(fs)
    return fs



col1 = ['Lãi gộp', 'op', 'EBT', 'Lãi/(lỗ) ròng trước thuế', 'Lãi/(lỗ) thuần sau thuế', 'Lợi nhuận của Cổ đông của Công ty mẹ', 'core_e','EBITDA']
col3 = ['Doanh số thuần', 'Lãi gộp', 'op', 'EBT', 'Lãi/(lỗ) ròng trước thuế', 'Lãi/(lỗ) thuần sau thuế', 'Lợi nhuận của Cổ đông của Công ty mẹ', 'core_e','EBITDA']
col2 = ['Doanh số thuần', 'Lãi gộp', 'op', 'EBT', 'Lãi/(lỗ) ròng trước thuế', 'Lãi/(lỗ) thuần sau thuế', 'Lợi nhuận của Cổ đông của Công ty mẹ', 'core_e','EBITDA', 
        'Doanh số thuần_4Q', 'Lãi gộp_4Q', 'op_4Q','Lãi/(lỗ) ròng trước thuế', 'Lãi/(lỗ) thuần sau thuế_4Q', 'Lợi nhuận của Cổ đông của Công ty mẹ_4Q', 'core_e_4Q','EBITDA_4Q']

def margin_func_Q(x):
    for i in col1:
        margin_col = i + "m"
        x = x.with_columns([(pl.col(i)/(1+pl.col('Doanh thu thuần'))).alias(margin_col)])
    return x

def ttm(x):
    for i in col1:
        ttm = i+"_4Q"
        x = x.with_columns(pl.col(i).rolling_sum(window_size=4).alias(ttm))
    return x

def g_func_Q(x):
    for i in col3:
        g_col = "g_"+i
        x = x.with_columns((pl.col(i).pct_change(n=3)).alias(g_col))
    return x

def get_fs_Q(ticker):
    bs = financial_report(ticker,'BalanceSheet','Quarterly')
    pl = financial_report(ticker,'IncomeStatement','Quarterly')
    cf = financial_report(ticker,'CashFlow','Quarterly')
    cf = cf.rename(columns={'Unnamed: 0': 'CHỈ TIÊU'})
    cf.set_index('CHỈ TIÊU', inplace=True)
    cf2 = pd.DataFrame(index = {'Khấu hao TSCĐ':'CHỈ TIÊU'}, columns = cf.columns).fillna(0)
    if cf2.index[0] in cf.index:
        pass
    else:
        cf = pd.concat([cf,cf2],axis=0)
    cf = cf.reset_index().rename(columns={'index': 'CHỈ TIÊU'})

    fs = pd.concat([bs,pl,cf])
    #delete all the column with NaN value > 40
    fs = fs.dropna(axis=1,thresh=40)
    fs = fs.T
    fs.columns = fs.iloc[0]
    fs = fs.iloc[1:,:]
    
    #Dropping rows if more than half of the values are zeros 
    # fs = fs.loc[fs.isna().sum(axis=1)<50]

    fs['year'] = fs.index.str[-4:].astype(int)
    fs['quarter'] = fs.index.str[1].astype(float)
    fs['dates'] = pd.PeriodIndex(year=fs["year"], quarter=fs["quarter"])
    fs['dates'] = fs['dates'].dt.to_timestamp(freq='Q')
    fs = fs.sort_values(by='dates')
    try:
        mc = get_mc_Q(ticker)
    except:
        mc = pd.DataFrame(columns=['mc','dates'])
    fs = fs.merge(mc[['mc','dates']], on='dates', how='left')
    return fs


def get_data_Q(ticker):
    x = get_fs_Q(ticker)
    x = x.loc[:,~x.columns.duplicated(keep='first')]
    x = pl.from_pandas(x)
    x = add_ratios(x)
    x = margin_func(x)
    x = ttm(x)
    x = g_func_Q(x)

    x = x.with_columns([(pl.col('dates').dt.strftime("%Y-%m")).alias('dates')])
    return x


import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle,Font, Color, Alignment, Border, Side, PatternFill
desktop = os.path.expanduser("~/Desktop")
os.chdir(desktop)
list_row = ['mc','Doanh số thuần',
    'Lãi gộp',
    'Chi phí bán hàng',
    'Chi phí quản lý doanh  nghiệp',
    'Thu nhập tài chính',
    'Chi phí tài chính',
    'Trong đó: Chi phí lãi vay',
    'Lãi/(lỗ) từ công ty liên doanh',
    'Thu nhập khác, ròng',
    'Lãi/(lỗ) ròng trước thuế',
    'Lãi/(lỗ) thuần sau thuế',
    'Lợi nhuận của Cổ đông của Công ty mẹ',
    'operating_EBITDA','EBITDA','op','core_e','fin_income','EBT', 
    
    'TÀI SẢN NGẮN HẠN','TÀI SẢN DÀI HẠN',
    'bs_cash','bs_ar','Hàng tồn kho, ròng',
    'bs_fa','Phải thu dài hạn','Tài sản dở dang dài hạn','Đầu tư dài hạn','other_asset',
    'TỔNG TÀI SẢN',
    
    'NỢ PHẢI TRẢ','Vay ngắn hạn','Vay dài hạn','bs_cust_pre','Phải trả người bán','other_lia',
    'VỐN CHỦ SỞ HỮU','Vốn góp','Thặng dư vốn cổ phần','Cổ phiếu Quỹ','Lãi chưa phân phối','Lợi ích cổ đông không kiểm soát','other_equity','netdebt',
    
    'Lưu chuyển tiền thuần từ các hoạt động sản xuất kinh doanh','Khấu hao TSCĐ',
    'Lưu chuyển tiền tệ ròng từ hoạt động đầu tư','Tiền mua tài sản cố định và các tài sản dài hạn khác','Tiền thu được từ thanh lý tài sản cố định','Cổ tức và tiền lãi nhận được',
    'Lưu chuyển tiền tệ từ hoạt động tài chính',
    'Cổ tức đã trả',
    'Tiền thu từ phát hành cổ phiếu và vốn góp','Chi trả cho việc mua lại, trả lại cổ phiếu',
    'Tiền thu được các khoản đi vay','Tiển trả các khoản đi vay','Tiền thanh toán vốn gốc đi thuê tài chính',
    'cf_div','cf_delta_debt','cf_khac',
    'ca/ta','de','tax_rate'
    ]



header = NamedStyle(name="header")
header.font = Font(bold=True, color="ffffff", size=12,name='Roboto')
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment( horizontal="center", vertical="center")
header.fill = PatternFill(fgColor= "14233c", fill_type="solid")

def formatting(ticker,data,worksheet):
    ticker = ticker.upper()
    wb = load_workbook(ticker+'.xlsx')
    ws1 = wb[worksheet]

    for r in dataframe_to_rows(data, index=True, header=True):
        ws1.append(r)
    ws1.sheet_view.showGridLines = False
    

    header_row = ws1[1]
    for cell in header_row:
        cell.style = header
    ws1.column_dimensions['A'].width = 40
    for cell in ws1['A']:
        cell.font = Font(size=10,name='Roboto')
    #divide by 10**9 except last 3 rows, and first column
    for col in ws1.iter_cols(min_row=2, max_row=ws1.max_row, min_col=2, max_col=ws1.max_column):
        for cell in col:
            if cell.row < ws1.max_row-2:
                try:
                    cell.value = cell.value/10**9
                except:
                    pass

    ws1.insert_rows(22),ws1.insert_rows(48),ws1.insert_rows(65)
    ws1['A2'] = 'INCOME STATEMENT'
    ws1['A22'] = 'BALANCE SHEET'
    ws1['A48'] = 'CASH FLOW'
    ws1['A65'] = 'RATIOS'
    for i in [2,22,48,65]:
        ws1.cell(row=i, column=1).style = 'header'

    #change number format for all columns except row 1
    for col in ws1.iter_cols(min_row=2, max_row=ws1.max_row, min_col=2, max_col=ws1.max_column):
        for cell in col:
            cell.number_format = '#,##0.0'
            cell.font = Font(size=10,name='Roboto')
            
    wb.save(ticker+'.xlsx')

def save_excel(ticker):
    ticker = ticker.upper()
    x = get_data_Y(ticker)
    x = x.to_pandas()
    x.to_excel(ticker+'.xlsx',index=False,sheet_name='Raw_data_Y')
    x.set_index('dates',inplace=True)
    x = x.T
    dataY = x.loc[list_row]

    y = get_data_Q(ticker)
    y = y.to_pandas()
    
    wb = load_workbook(ticker+'.xlsx')
    wsQ = wb.create_sheet('Raw_data_Q')
    for r in dataframe_to_rows(y, index=True, header=True):
        wsQ.append(r)
    
    y.set_index('dates',inplace=True)
    y = y.T
    dataQ = y.loc[list_row]
    
    ws1 = wb.create_sheet('Indicator_Y')
    ws2 = wb.create_sheet('Indicator_Q')
    wb.save(ticker+'.xlsx')
    formatting(ticker,dataY,'Indicator_Y')
    formatting(ticker,dataQ,'Indicator_Q')
#create input "Nhập ticker: " and run save_excel function
while True:
    ticker = input("Nhập ticker: ")
    try:
        save_excel(ticker)
        print("Đã tạo file excel cho " + ticker.upper())
    except Exception as e:
        print("Có lỗi xảy ra, vui lòng kiểm tra lại ticker")
        print(e)


